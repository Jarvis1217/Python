import re
import openpyxl
from textwrap import dedent

class ExcelDB:
    def __init__(self, filename='database.xlsx'):
        self.filename = filename
        try:
            self.wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            ws = self.wb.active
            ws.title = "db_init"

        self.save()

    def save(self):
        self.wb.save(self.filename)

    def execute(self, sql: str):
        sql = sql.strip().rstrip(';')
        cmd = sql.split()[0].upper()
        if cmd == 'CREATE':
            return self._create_table(sql)
        elif cmd == 'INSERT':
            return self._insert(sql)
        elif cmd == 'SELECT':
            return self._select(sql)
        elif cmd == 'UPDATE':
            return self._update(sql)
        elif cmd == 'DELETE':
            return self._delete(sql)
        else:
            raise ValueError(f"Unsupported command: {cmd}")

    def _create_table(self, sql: str):
        # CREATE TABLE table_name (col1, col2, ...)
        pattern = r"CREATE\s+TABLE\s+(\w+)\s*\((.+)\)"
        m = re.match(pattern, sql, re.IGNORECASE)
        if not m:
            raise ValueError("Invalid CREATE TABLE syntax")
        table, cols = m.group(1), m.group(2)
        cols = [c.strip() for c in cols.split(',')]
        if table in self.wb.sheetnames:
            raise ValueError(f"Table {table} already exists")
        ws = self.wb.create_sheet(table)
        for idx, col in enumerate(cols, start=1):
            ws.cell(row=1, column=idx, value=col)
        self.save()
        return f"Table {table} created with columns {cols}"

    def _insert(self, sql: str):
        # INSERT INTO table_name (col1, col2) VALUES (val1, val2)
        pattern = r"INSERT\s+INTO\s+(\w+)\s*\((.+)\)\s*VALUES\s*\((.+)\)"
        m = re.match(pattern, sql, re.IGNORECASE)
        if not m:
            raise ValueError("Invalid INSERT syntax")
        table, cols, vals = m.group(1), m.group(2), m.group(3)
        cols = [c.strip() for c in cols.split(',')]
        vals = [v.strip().strip("'\"") for v in re.split(r",(?=(?:[^']*'[^']*')*[^']*$)", vals)]
        ws = self.wb[table]
        header = [cell.value for cell in ws[1]]
        row = [''] * len(header)
        for c, v in zip(cols, vals):
            idx = header.index(c)
            row[idx] = v
        ws.append(row)
        self.save()
        return f"Inserted into {table}: {dict(zip(cols, vals))}"

    def _select(self, sql: str):
        # SELECT col1, col2 FROM table_name [WHERE col = 'value']
        pattern = r"SELECT\s+(.+)\s+FROM\s+(\w+)(?:\s+WHERE\s+(.+))?"
        m = re.match(pattern, sql, re.IGNORECASE)
        if not m:
            raise ValueError("Invalid SELECT syntax")
        cols, table, where = m.group(1), m.group(2), m.group(3)
        ws = self.wb[table]
        header = [cell.value for cell in ws[1]]
        if cols.strip() == '*':
            sel_idxs = list(range(len(header)))
        else:
            sel_cols = [c.strip() for c in cols.split(',')]
            sel_idxs = [header.index(c) for c in sel_cols]
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if where:
                wcol, wval = [part.strip() for part in where.split('=')]
                wval = wval.strip("'\" ")
                if str(row[header.index(wcol)]) != wval:
                    continue
            results.append([row[i] for i in sel_idxs])

        cols_out = header if cols.strip()=='*' else [header[i] for i in sel_idxs]
        return [dict(zip(cols_out, r)) for r in results]

    def _update(self, sql: str):
        # UPDATE table_name SET col = 'value' [WHERE col = 'value']
        pattern = r"UPDATE\s+(\w+)\s+SET\s+(.*?)\s*(?:WHERE\s+(.*?))?\s*;?\s*$"
        m = re.match(pattern, sql, re.IGNORECASE)
        if not m:
            raise ValueError("Invalid UPDATE syntax")
        table, set_clause, where = m.group(1), m.group(2), m.group(3)
        ws = self.wb[table]
        header = [cell.value for cell in ws[1]]
        set_pairs = [p.strip() for p in set_clause.split(',')]
        updates = {}
        for pair in set_pairs:
            col, val = [x.strip() for x in pair.split('=')]
            updates[col] = val.strip("'\"")
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if where:
                wcol, wval = [part.strip() for part in where.split('=')]
                wval = wval.strip("'\"")
                if str(ws.cell(row=row_idx, column=header.index(wcol)+1).value) != wval:
                    continue
            for col, newval in updates.items():
                ws.cell(row=row_idx, column=header.index(col)+1, value=newval)
            count += 1
        self.save()
        return f"Updated {count} rows in {table}"

    def _delete(self, sql: str):
        # DELETE FROM table_name WHERE col = 'value'
        pattern = r"DELETE\s+FROM\s+(\w+)(?:\s+WHERE\s+(.+))?"
        m = re.match(pattern, sql, re.IGNORECASE)
        if not m:
            raise ValueError("Invalid DELETE syntax")
        table, where = m.group(1), m.group(2)
        ws = self.wb[table]
        header = [cell.value for cell in ws[1]]
        to_delete = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if where:
                wcol, wval = [part.strip() for part in where.split('=')]
                wval = wval.strip("'\"")
                if str(row[header.index(wcol)]) == wval:
                    to_delete.append(row_idx)
            else:
                to_delete.append(row_idx)
        for idx in reversed(to_delete):
            ws.delete_rows(idx)
        self.save()
        return f"Deleted {len(to_delete)} rows from {table}"

if __name__ == '__main__':
    db = ExcelDB()

    # 测试
    # print(db.execute("CREATE TABLE users (id, name, email);"))
    # print(db.execute("INSERT INTO users (id, name, email) VALUES (1, 'Alice', 'alice@example.com');"))
    # print(db.execute("INSERT INTO users (id, name, email) VALUES (2, 'Bob', 'bob@example.com');"))
    # print(db.execute("SELECT * FROM users;"))
    # print(db.execute("UPDATE users SET email = 'alice@new.com' WHERE id = 1;"))
    # print(db.execute("DELETE FROM users;"))

    while True:
        sql_string = input(">>> ")

        if sql_string.casefold() == 'help'.casefold():
            print(dedent("""
                    CREATE TABLE table_name (col1, col2, ...);
                    INSERT INTO table_name (col1, col2) VALUES (val1, val2);
                    SELECT col1, col2 FROM table_name [WHERE col = 'value'];
                    UPDATE table_name SET col = 'value' [WHERE col = 'value'];
                    DELETE FROM table_name WHERE col = 'value';
                """))
            continue
        
        try:
            print(db.execute(sql_string))
        except Exception as e:
            print(e)
