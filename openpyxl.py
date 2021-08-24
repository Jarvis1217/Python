from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook("./是否集团客户(1).xlsx")
    ws = wb["Sheet1"]

    z, a, b = "", "", ""

    for i in range(ws.min_row + 1, ws.max_row + 1):
        for j in range(ws.min_column, 6):
            if j == 2:
                z = ws.cell(i, j).value
            if j == 4:
                a = ws.cell(i, j).value
            if j == 5:
                b = ws.cell(i, j).value if ws.cell(i, j).value is not None else "null"

        if b == "null":
            print(f"update hls_bp_master set group_customers = '{a.strip()}',group_membership = {b.strip()} where bp_code = '{z.strip()}';")
        else:
            print(f"update hls_bp_master set group_customers = '{a.strip()}',group_membership = '{b.strip()}' where bp_code = '{z.strip()}';")
