import pymysql
from flask import Flask, render_template, request, redirect, flash, url_for
from openpyxl import load_workbook

app = Flask(__name__, template_folder='../templates', static_folder="../static")
app.config["SECRET_KEY"] = "this_is_a_secretKey"
wb = load_workbook("../Data.xlsx")


@app.route("/")
def index():
    return redirect(url_for("userLogin"))


@app.route("/Login", methods=["GET", "POST"])
def userLogin():
    if request.method == "GET":
        return render_template("userLogin.html")
    if request.method == "POST":
        user_info = request.form.to_dict()
        name = user_info.get("userName")
        passwd = user_info.get("userPass")
        db = pymysql.connect(host="localhost", port=3306, user="root", password="root", database="library",
                             charset="utf8")
        cur = db.cursor()
        cur.execute("select * from userInfo where userName='%s'" % name)
        data = cur.fetchone()
        db.close()
        if data is not None:
            data = list(data)
            if name == data[1] and passwd == data[2]:
                return redirect(url_for("Show"))
            else:
                flash("登录信息有误,请核对后登录!")
                return redirect(url_for("userLogin"))
        else:
            flash("登录信息有误,请核对后登录!")
            return redirect(url_for("userLogin"))


@app.route("/Register", methods=["GET", "POST"])
def userRegister():
    if request.method == "GET":
        return render_template("userRegister.html")
    if request.method == "POST":
        user_info = request.form.to_dict()
        name = user_info.get("userName")
        passwd = user_info.get("userPass")
        againpass = user_info.get("againPass")
        if passwd == againpass:
            db = pymysql.connect(host="localhost", port=3306, user="root", password="root", database="library",
                                 charset="utf8")
            cur = db.cursor()
            try:
                cur.execute("INSERT INTO userInfo (userName,userPasswd) VALUES ('%s','%s')" % (name, passwd))
                db.commit()
            except Exception as e:
                print(e)
                db.rollback()
            finally:
                db.close()
            flash("注册成功,请返回登录!")
            return redirect(url_for("userLogin"))
        else:
            flash("两次输入密码不一致,请核对!")
            return redirect(url_for("userRegister"))


@app.route("/Show")
def Show():
    datas = read_excel()
    return render_template("Show.html", datas=datas)


@app.route("/Buy", methods=["GET", "POST"])
def Buy():
    if request.method == "GET":
        return render_template("Buy.html")
    if request.method == "POST":
        user_info = request.form.to_dict()
        id = user_info.get("id")
        getAddress = user_info.get("getAddress")
        telNumber = user_info.get("telNumber")
        ws = wb["buyInfo"]
        ws.append([getAddress, telNumber])
        ws = wb["library"]
        ws.cell(int(id)+1, 5).value = ws.cell(int(id)+1, 5).value - 1
        wb.save("../Data.xlsx")
        return redirect(url_for("Show"))
    pass


def read_excel():
    sheet = wb["library"]
    max_row = sheet.max_row
    max_column = sheet.max_column
    data = []
    for r in range(2, max_row + 1):
        subdata = {}
        for c in range(1, max_column + 1):
            key = sheet.cell(1, c).value
            subdata[key] = sheet.cell(r, c).value
        data.append(subdata)
    return data


if __name__ == '__main__':
    app.run()